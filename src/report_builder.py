import logging
from datetime import date
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

from utils import month_label, pct

logger = logging.getLogger(__name__)

# ── Color palette ───────────────────────────────────────────────────────────────
FILL_WHITE = PatternFill(fill_type=None)
FILL_AMBER = PatternFill("solid", fgColor="FFC000")
FILL_RED_LIGHT = PatternFill("solid", fgColor="FFCCCC")
FILL_GREEN = PatternFill("solid", fgColor="E2EFDA")
FILL_HEADER_DARK = PatternFill("solid", fgColor="1F4E79")
FILL_HEADER_MID = PatternFill("solid", fgColor="2E75B6")

FONT_TITLE = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
FONT_HEADER = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
FONT_BODY = Font(name="Calibri", size=11)
FONT_BODY_BOLD = Font(name="Calibri", bold=True, size=11)
FONT_ERROR = Font(name="Calibri", size=11, color="FF0000", bold=True)
FONT_NOTE = Font(name="Calibri", size=10, italic=True, color="595959")

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

THIN = Side(style="thin", color="BFBFBF")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

SOURCE_AUTOMATED = "HubSpot CRM API"
SOURCE_MANUAL = "⚠ Manual entry required"
SOURCE_ADS = "HubSpot Ads API"
SOURCE_DASHBOARD = "HubSpot Dashboard (manual)"
SOURCE_6SENSE = "⚠ Manual entry required — pull from 6Sense platform directly"
SOURCE_FORMULA = "Calculated"


def _apply_col_widths(ws, widths: dict[int, int]):
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width


def _title_row(ws, title: str, ncols: int = 3):
    ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
    cell = ws["A1"]
    cell.value = title
    cell.fill = FILL_HEADER_DARK
    cell.font = FONT_TITLE
    cell.alignment = ALIGN_CENTER
    ws.row_dimensions[1].height = 28


def _note_row(ws, note: str, ncols: int = 3):
    ws.merge_cells(f"A2:{get_column_letter(ncols)}2")
    cell = ws["A2"]
    cell.value = note
    cell.font = FONT_NOTE
    cell.alignment = ALIGN_LEFT
    ws.row_dimensions[2].height = 16


def _header_row(ws, col_label: str, row: int = 3, ncols: int = 3):
    headers = ["Metric", col_label, "Source"]
    for i, h in enumerate(headers[:ncols], 1):
        cell = ws.cell(row=row, column=i, value=h)
        cell.fill = FILL_HEADER_MID
        cell.font = FONT_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER
    ws.row_dimensions[row].height = 20


def _data_row(ws, row: int, metric: str, value: Any, source: str, fill=None, error: bool = False):
    cells = [
        ws.cell(row=row, column=1, value=metric),
        ws.cell(row=row, column=2, value=value),
        ws.cell(row=row, column=3, value=source),
    ]
    for cell in cells:
        cell.alignment = ALIGN_LEFT
        cell.border = THIN_BORDER
        if fill:
            cell.fill = fill
        if error:
            cell.font = FONT_ERROR
        else:
            cell.font = FONT_BODY
    cells[0].font = FONT_BODY_BOLD if not error else FONT_ERROR
    ws.row_dimensions[row].height = 18


def _val(data: dict, key: str, fallback="API ERROR") -> Any:
    if "error" in data:
        return fallback
    return data.get(key, fallback)


# ── Sheet 1: Lead / MQL Metrics ────────────────────────────────────────────────

def build_sheet1(wb: Workbook, data: dict, config: dict, year: int, month: int):
    ws = wb.create_sheet("Lead & MQL Metrics")
    label = month_label(year, month)
    goals = config["goals"]

    _title_row(ws, "BillingPlatform — Lead & MQL Metrics")
    _note_row(ws, "White = HubSpot API  |  Amber = Manual entry required")
    _header_row(ws, label)

    def row(r, metric, value, source, fill=FILL_WHITE, error=False):
        _data_row(ws, r, metric, value, source, fill=fill, error=error)

    error = "error" in data

    leads_actual = _val(data, "leads", 0) if not error else "API ERROR"
    leads_goal = goals["leads_goal"]
    mqls_actual = _val(data, "mqls", 0) if not error else "API ERROR"
    mqls_goal = goals["mqls_goal"]
    sals_actual = _val(data, "sals_from_mqls", 0) if not error else "API ERROR"
    sals_goal = goals["sals_goal"]
    meetings_actual = _val(data, "meetings", 0) if not error else "API ERROR"
    meetings_goal = goals["meetings_goal"]

    rows = [
        (4,  "Product Pipeline Created ($)",    _val(data, "pipeline_created"), SOURCE_AUTOMATED),
        (5,  "Leads (Actual)",                  leads_actual,                  SOURCE_AUTOMATED),
        (6,  "Leads (Goal)",                    leads_goal,                    "config.yaml"),
        (7,  "Leads (Goal Attainment %)",       pct(leads_actual, leads_goal) if not error else "N/A", SOURCE_FORMULA),
        (8,  "Total MQLs (Actual)",             mqls_actual,                   SOURCE_AUTOMATED),
        (9,  "MQLs (Goal)",                     mqls_goal,                     "config.yaml"),
        (10, "MQLs (Goal Attainment %)",        pct(mqls_actual, mqls_goal) if not error else "N/A", SOURCE_FORMULA),
        (11, "6QA Accounts",                    _val(data, "six_qa_accounts"),  SOURCE_AUTOMATED),
        (12, "SALs from MQLs (Actual)",         sals_actual,                   SOURCE_AUTOMATED),
        (13, "SALs (Goal)",                     sals_goal,                     "config.yaml"),
        (14, "SALs (Goal Attainment %)",        pct(sals_actual, sals_goal) if not error else "N/A", SOURCE_FORMULA),
        (15, "Meetings (Actual)",               meetings_actual,               SOURCE_AUTOMATED),
        (16, "Meetings (Goal)",                 meetings_goal,                 "config.yaml"),
        (17, "Meetings (Goal Attainment %)",    pct(meetings_actual, meetings_goal) if not error else "N/A", SOURCE_FORMULA),
    ]

    formula_rows = {7, 10, 14, 17}
    goal_rows = {6, 9, 13, 16}

    for r, metric, value, source in rows:
        is_error = isinstance(value, str) and value == "API ERROR"
        fill = FILL_WHITE
        row(r, metric, value, source, fill=fill, error=is_error)

    _apply_col_widths(ws, {1: 38, 2: 18, 3: 38})
    logger.info("Sheet 1 built")


# ── Sheet 2: Email Metrics ──────────────────────────────────────────────────────

def build_sheet2(wb: Workbook, data: dict, year: int, month: int):
    ws = wb.create_sheet("Email Metrics")
    label = month_label(year, month)

    _title_row(ws, "BillingPlatform — Email Metrics")
    _note_row(ws, "White = HubSpot API  |  Amber = Manual entry required")
    _header_row(ws, label)

    error = "error" in data

    rows = [
        (4,  "Emails Sent",                     _val(data, "emails_sent"),          SOURCE_AUTOMATED,  FILL_WHITE),
        (5,  "Unique Contacts Receiving Emails", None,                               SOURCE_MANUAL,     FILL_AMBER),
        (6,  "Conversion Rate",                 None,                               SOURCE_MANUAL,     FILL_AMBER),
        (7,  "Delivery Rate",                   _val(data, "delivery_rate"),         SOURCE_AUTOMATED,  FILL_WHITE),
        (8,  "Open Rate",                       _val(data, "open_rate"),             SOURCE_AUTOMATED,  FILL_WHITE),
        (9,  "Click To Open Rate",              None,                               SOURCE_MANUAL,     FILL_AMBER),
        (10, "Click Thru Rate",                 _val(data, "click_through_rate"),    SOURCE_AUTOMATED,  FILL_WHITE),
    ]

    for r, metric, value, source, fill in rows:
        is_error = isinstance(value, str) and value == "API ERROR"
        _data_row(ws, r, metric, value, source, fill=fill, error=is_error)

    _apply_col_widths(ws, {1: 38, 2: 18, 3: 38})
    logger.info("Sheet 2 built")


# ── Sheet 3: LinkedIn Metrics ───────────────────────────────────────────────────

def build_sheet3(wb: Workbook, data: dict, year: int, month: int):
    ws = wb.create_sheet("LinkedIn Metrics")
    label = month_label(year, month)

    _title_row(ws, "BillingPlatform — LinkedIn Ad Metrics")
    _note_row(ws, "Light red = HubSpot Ads API  |  Green = HubSpot Dashboard (manual)  |  Amber = Manual entry required")
    _header_row(ws, label)

    error = "error" in data

    brand_spend = _val(data, "brand_spend")
    brand_mqls = _val(data, "brand_mqls")
    abm_spend = _val(data, "abm_spend")
    abm_mqls = _val(data, "abm_mqls")
    total_spend = _val(data, "total_paid_social_spend")
    total_cpl = _val(data, "total_cpl")

    rows = [
        (4,  "Brand Spend ($)",                         brand_spend,               SOURCE_ADS,       FILL_RED_LIGHT),
        (5,  "MQLs (Brand)",                            brand_mqls,                SOURCE_ADS,       FILL_RED_LIGHT),
        (6,  "Avg CTR",                                 _val(data, "avg_ctr"),     SOURCE_ADS,       FILL_RED_LIGHT),
        (7,  "Clicks",                                  _val(data, "clicks"),      SOURCE_ADS,       FILL_RED_LIGHT),
        (8,  "CPL ($)",                                 _val(data, "cpl"),         SOURCE_FORMULA,   FILL_RED_LIGHT),
        (9,  "Impressions",                             _val(data, "impressions"), SOURCE_ADS,       FILL_RED_LIGHT),
        (10, "Avg CPM ($)",                             _val(data, "avg_cpm"),     SOURCE_ADS,       FILL_RED_LIGHT),
        (11, "ABM Spend ($)",                           abm_spend,                 SOURCE_ADS,       FILL_RED_LIGHT),
        (12, "Total ABM MQLs",                         abm_mqls,                  SOURCE_ADS,       FILL_RED_LIGHT),
        (13, "Total ABM Qualified/Converted MQLs",     None,                      SOURCE_DASHBOARD, FILL_GREEN),
        (14, "Qualified MQL Production Rate",           "N/A — manual row 13",     SOURCE_FORMULA,   FILL_GREEN),
        (15, "Total Paid Social Spend ($)",             total_spend,               SOURCE_FORMULA,   FILL_RED_LIGHT),
        (16, "Total CPL ($)",                          total_cpl,                  SOURCE_FORMULA,   FILL_RED_LIGHT),
        (17, "Total CPL for Converted Leads Only ($)", "N/A — manual row 13",     SOURCE_FORMULA,   FILL_GREEN),
    ]

    for r, metric, value, source, fill in rows:
        is_error = isinstance(value, str) and value == "API ERROR"
        _data_row(ws, r, metric, value, source, fill=fill, error=is_error)

    _apply_col_widths(ws, {1: 42, 2: 18, 3: 42})
    logger.info("Sheet 3 built")


# ── Sheet 4: Google Metrics ─────────────────────────────────────────────────────

def build_sheet4(wb: Workbook, data: dict, year: int, month: int):
    ws = wb.create_sheet("Google Metrics")
    label = month_label(year, month)

    _title_row(ws, "BillingPlatform — Google Ad Metrics")
    _note_row(ws, "Light red = HubSpot Ads API  |  Green = HubSpot Dashboard (manual)")
    _header_row(ws, label)

    total_cost = _val(data, "total_cost")
    total_mqls = _val(data, "total_mqls")

    rows = [
        (4,  "Total Cost ($)",                          total_cost,                        SOURCE_ADS,       FILL_RED_LIGHT),
        (5,  "Total MQLs",                              total_mqls,                        SOURCE_ADS,       FILL_RED_LIGHT),
        (6,  "Cost Per Lead ($)",                       _val(data, "cost_per_lead"),       SOURCE_FORMULA,   FILL_RED_LIGHT),
        (7,  "Conversion Rate",                         _val(data, "conversion_rate"),     SOURCE_ADS,       FILL_RED_LIGHT),
        (8,  "Clicks",                                  _val(data, "clicks"),              SOURCE_ADS,       FILL_RED_LIGHT),
        (9,  "CTR",                                     _val(data, "ctr"),                 SOURCE_ADS,       FILL_RED_LIGHT),
        (10, "Avg Cost Per Click ($)",                  _val(data, "avg_cpc"),             SOURCE_ADS,       FILL_RED_LIGHT),
        (11, "Impressions",                             _val(data, "impressions"),         SOURCE_ADS,       FILL_RED_LIGHT),
        (12, "Converted Leads / Qualified MQLs",       None,                              SOURCE_DASHBOARD, FILL_GREEN),
        (13, "Qualified MQL Production Rate",           "N/A — manual row 12",             SOURCE_FORMULA,   FILL_GREEN),
        (14, "Total CPL ($)",                           _val(data, "total_cpl"),           SOURCE_FORMULA,   FILL_RED_LIGHT),
        (15, "Total CPL for Converted Leads Only ($)", "N/A — manual row 12",             SOURCE_FORMULA,   FILL_GREEN),
    ]

    for r, metric, value, source, fill in rows:
        is_error = isinstance(value, str) and value == "API ERROR"
        _data_row(ws, r, metric, value, source, fill=fill, error=is_error)

    _apply_col_widths(ws, {1: 42, 2: 18, 3: 42})
    logger.info("Sheet 4 built")


# ── Sheet 5: 6Sense Metrics ─────────────────────────────────────────────────────

def build_sheet5(wb: Workbook, year: int, month: int):
    ws = wb.create_sheet("6Sense Metrics")
    label = month_label(year, month)

    _title_row(ws, "BillingPlatform — 6Sense Metrics")
    _note_row(ws, "All rows require manual entry — 6Sense does not have a HubSpot API integration")
    _header_row(ws, label)

    metrics = [
        "Accounts Targeted (YTD)",
        "Accounts Reached (YTD)",
        "Accounts Engaged (YTD)",
        "Accounts Targeted (MTD)",
        "Accounts Reached (MTD)",
        "Accounts Engaged (MTD)",
        "Impressions (MTD)",
        "Total Cost (MTD)",
        "# of 6QAs",
        "% 6QAs Worked by Sales",
    ]

    for i, metric in enumerate(metrics, start=4):
        _data_row(ws, i, metric, None, SOURCE_6SENSE, fill=FILL_AMBER)

    _apply_col_widths(ws, {1: 38, 2: 18, 3: 55})
    logger.info("Sheet 5 built")


# ── Main entry point ────────────────────────────────────────────────────────────

def build_report(
    sheet1_data: dict,
    sheet2_data: dict,
    sheet3_data: dict,
    sheet4_data: dict,
    config: dict,
    year: int,
    month: int,
    output_path: str,
):
    wb = Workbook()
    # Remove default empty sheet
    wb.remove(wb.active)

    build_sheet1(wb, sheet1_data, config, year, month)
    build_sheet2(wb, sheet2_data, year, month)
    build_sheet3(wb, sheet3_data, year, month)
    build_sheet4(wb, sheet4_data, year, month)
    build_sheet5(wb, year, month)

    wb.save(output_path)
    logger.info("Report saved to %s", output_path)
